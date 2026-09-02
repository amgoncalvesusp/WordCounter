"""Export results to XLSX using a dynamic, analyzer-driven column schema."""

import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError

from .analysis import ColumnSpec, build_column_specs, build_default_analyzers

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
TERM_HEADER_FILL = PatternFill(
    start_color="7B5CB8", end_color="7B5CB8", fill_type="solid"
)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
INVALID_EXCEL_XML_CHARS_RE = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F]+"
)


class ExcelExportError(RuntimeError):
    """Raised when a value cannot be safely written to an XLSX worksheet."""


def _coerce_excel_text(value):
    """Convert byte-oriented text to Unicode before sanitizing it."""
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, bytearray):
        return bytes(value).decode("utf-8", errors="replace")
    return value


def _sanitize_excel_value(value):
    """Return a value safe to write to an XLSX cell.

    Only XML-invalid control characters are replaced.
    Valid Unicode, accents, superscripts, tabs and line breaks are preserved.
    """
    value = _coerce_excel_text(value)
    if not isinstance(value, str):
        return value

    return INVALID_EXCEL_XML_CHARS_RE.sub(" ", value)


def _invalid_excel_xml_codepoints(value) -> List[str]:
    """Return invalid XML code points without retaining document content."""
    text = _coerce_excel_text(value)
    if not isinstance(text, str):
        return []

    return sorted(
        {
            f"U+{ord(character):04X}"
            for character in text
            if INVALID_EXCEL_XML_CHARS_RE.fullmatch(character)
        }
    )


def _write_cell(ws, row: int, column: int, value):
    """Write a sanitized value and return the created cell."""
    try:
        cell = ws.cell(
            row=row,
            column=column,
            value=_sanitize_excel_value(value),
        )
    except IllegalCharacterError as exc:
        codepoints = _invalid_excel_xml_codepoints(value)
        detail = ", ".join(codepoints) or "não identificados"
        coordinate = f"{get_column_letter(column)}{row}"
        raise ExcelExportError(
            f"Não foi possível exportar a aba '{ws.title}', célula {coordinate}. "
            f"Caracteres XML inválidos: {detail}."
        ) from exc

    if cell.data_type == "f":
        # Text extracted from PDFs can start with "=", which openpyxl would write
        # as a formula; Excel then rejects it ("Registros Removidos: Fórmula").
        # Everything we export is data, never a formula, so force the text type.
        cell.data_type = "s"
    return cell


def export_to_xlsx(
    results: List[Dict],
    output_path: str,
    column_specs: Optional[List[ColumnSpec]] = None,
) -> None:
    """Write results to XLSX.

    ``column_specs`` is the full ordered output schema (see
    :func:`src.core.analysis.build_column_specs`). When omitted it is derived
    from the term columns present in the results, preserving backward
    compatibility with callers that do not pass an explicit schema.
    """
    if column_specs is None:
        column_specs = _infer_column_specs(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contagem de Palavras"

    for col_idx, spec in enumerate(column_specs, start=1):
        cell = _write_cell(ws, row=1, column=col_idx, value=spec.label)
        cell.fill = TERM_HEADER_FILL if spec.group == "term" else HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = spec.width

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "C2"

    for row_idx, result in enumerate(results, start=2):
        enriched = {"doc_id": row_idx - 1, **result}
        for col_idx, spec in enumerate(column_specs, start=1):
            value = enriched.get(spec.key, "")
            cell = _write_cell(ws, row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _write_excluded_sheet(wb, results)
    _write_sentiment_sheet(wb, results)
    _write_keyword_sheet(wb, results)
    _write_kwic_sheet(wb, results)
    _write_climate_policy_sheets(wb, results)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _infer_column_specs(results: List[Dict]) -> List[ColumnSpec]:
    """Reconstruct a schema from the analyzer outputs present in the results."""
    seen = set()
    terms: List[tuple] = []
    for r in results:
        for label in r.get("term_results", {}).keys():
            if label not in seen:
                seen.add(label)
                exact = label.startswith('"') and label.endswith('"')
                terms.append((label.strip('"') if exact else label, exact))
    detect_sentiment = any("sent_n_sentencas" in r for r in results)
    detect_textmetrics = any("lex_ttr" in r for r in results)
    detect_climate_policy = any("climate_policy_profile" in r for r in results)
    analyzers = build_default_analyzers(
        terms,
        detect_sentiment=detect_sentiment,
        detect_textmetrics=detect_textmetrics,
        detect_climate_policy=detect_climate_policy,
    )
    return build_column_specs(analyzers)


def _write_excluded_sheet(wb: Workbook, results: List[Dict]) -> None:
    ws2 = wb.create_sheet("Páginas Excluídas")
    headers2 = [
        "Nº Doc.",
        "Arquivo",
        "Página",
        "Motivo da Exclusão",
        "Palavras na Página",
    ]
    for col_idx, label in enumerate(headers2, start=1):
        cell = _write_cell(ws2, row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 20
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for page in result.get("excluded_pages", []):
            _write_cell(ws2, row=detail_row, column=1, value=doc_idx)
            _write_cell(ws2, row=detail_row, column=2, value=result["filename"])
            _write_cell(ws2, row=detail_row, column=3, value=page["page_number"])
            _write_cell(
                ws2,
                row=detail_row,
                column=4,
                value=page["exclusion_reason"],
            )
            _write_cell(ws2, row=detail_row, column=5, value=page["word_count"])
            for col in range(1, 6):
                ws2.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws2.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _write_sentiment_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Per-sentence sentiment detail — the registry units for content analysis.

    Each row is one scored sentence with its page, compound valence and class,
    so aggregates on the main sheet are fully traceable back to the source text
    (supports Bardin's content analysis and Aguiar & Ozella's meaning nuclei).
    """
    has_detail = any(r.get("sentiment_sentences") for r in results)
    if not has_detail:
        return

    ws = wb.create_sheet("Sentimento (Sentenças)")
    headers = ["Nº Doc.", "Arquivo", "Página", "Sentença", "Compound", "Classe"]
    for col_idx, label in enumerate(headers, start=1):
        cell = _write_cell(ws, row=1, column=col_idx, value=label)
        cell.fill = TERM_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 90
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for sent in result.get("sentiment_sentences", []):
            _write_cell(ws, row=detail_row, column=1, value=doc_idx)
            _write_cell(ws, row=detail_row, column=2, value=result["filename"])
            _write_cell(ws, row=detail_row, column=3, value=sent["page"])
            _write_cell(ws, row=detail_row, column=4, value=sent["text"])
            _write_cell(ws, row=detail_row, column=5, value=sent["compound"])
            _write_cell(ws, row=detail_row, column=6, value=sent["classe"])
            ws.cell(row=detail_row, column=4).alignment = Alignment(
                vertical="center", wrap_text=True
            )
            for col in range(1, 7):
                ws.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _write_keyword_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Ranked keyword frequencies per document — registry units for content analysis."""
    has_detail = any(r.get("keyword_freq") for r in results)
    if not has_detail:
        return

    ws = wb.create_sheet("Frequência de Palavras")
    headers = ["Nº Doc.", "Arquivo", "Palavra", "Frequência"]
    for col_idx, label in enumerate(headers, start=1):
        cell = _write_cell(ws, row=1, column=col_idx, value=label)
        cell.fill = TERM_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for word, count in result.get("keyword_freq", []):
            _write_cell(ws, row=detail_row, column=1, value=doc_idx)
            _write_cell(ws, row=detail_row, column=2, value=result["filename"])
            _write_cell(ws, row=detail_row, column=3, value=word)
            _write_cell(ws, row=detail_row, column=4, value=count)
            for col in range(1, 5):
                ws.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _write_kwic_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Keyword-in-context concordance — the qualitative context units (Bardin)."""
    has_detail = any(r.get("kwic") for r in results)
    if not has_detail:
        return

    ws = wb.create_sheet("Concordância (KWIC)")
    headers = [
        "Nº Doc.",
        "Arquivo",
        "Página",
        "Termo",
        "Contexto à esquerda",
        "Ocorrência",
        "Contexto à direita",
    ]
    for col_idx, label in enumerate(headers, start=1):
        cell = _write_cell(ws, row=1, column=col_idx, value=label)
        cell.fill = TERM_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    widths = [8, 28, 8, 18, 50, 18, 50]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for line in result.get("kwic", []):
            _write_cell(ws, row=detail_row, column=1, value=doc_idx)
            _write_cell(ws, row=detail_row, column=2, value=result["filename"])
            _write_cell(ws, row=detail_row, column=3, value=line["page"])
            _write_cell(ws, row=detail_row, column=4, value=line["term"])
            _write_cell(
                ws, row=detail_row, column=5, value=line["left"]
            ).alignment = right_align
            _write_cell(
                ws, row=detail_row, column=6, value=line["keyword"]
            ).alignment = center_align
            _write_cell(ws, row=detail_row, column=7, value=line["right"])
            for col in range(1, 8):
                ws.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _write_climate_policy_sheets(wb: Workbook, results: List[Dict]) -> None:
    """Write climate-policy matrix, evidence and gaps when available."""
    if not any(r.get("climate_policy_profile") for r in results):
        return

    _write_climate_matrix_sheet(wb, results)
    _write_climate_evidence_sheet(wb, results)
    _write_climate_gaps_sheet(wb, results)


def _style_header(ws, headers: List[str], fill=TERM_HEADER_FILL) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = _write_cell(ws, row=1, column=col_idx, value=label)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _style_detail_row(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER
        if row % 2 == 0:
            ws.cell(row=row, column=col).fill = ALT_ROW_FILL


def _kind_label(kind: str) -> str:
    return {"sector": "Setor", "instrument": "Instrumento"}.get(kind, kind)


def _write_climate_matrix_sheet(wb: Workbook, results: List[Dict]) -> None:
    ws = wb.create_sheet("Politica Climatica")
    headers = [
        "N Doc.",
        "Arquivo",
        "Tipo",
        "Categoria",
        "Status",
        "Evidencias diretas",
        "Evidencias indiretas",
        "Evidencias totais",
    ]
    _style_header(ws, headers)
    widths = [8, 32, 14, 34, 18, 18, 20, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        profile = result.get("climate_policy_profile") or {}
        for kind in ("sector", "instrument"):
            items = (profile.get(kind) or {}).get("items", {})
            for data in items.values():
                _write_cell(ws, row=detail_row, column=1, value=doc_idx)
                _write_cell(ws, row=detail_row, column=2, value=result["filename"])
                _write_cell(ws, row=detail_row, column=3, value=_kind_label(kind))
                _write_cell(ws, row=detail_row, column=4, value=data["label"])
                _write_cell(ws, row=detail_row, column=5, value=data["status"])
                _write_cell(ws, row=detail_row, column=6, value=data["direct_hits"])
                _write_cell(ws, row=detail_row, column=7, value=data["indirect_hits"])
                _write_cell(ws, row=detail_row, column=8, value=data["total_hits"])
                _style_detail_row(ws, detail_row, len(headers))
                detail_row += 1


def _write_climate_evidence_sheet(wb: Workbook, results: List[Dict]) -> None:
    ws = wb.create_sheet("Evidencias Climaticas")
    headers = [
        "N Doc.",
        "Arquivo",
        "Pagina",
        "Tipo",
        "Categoria",
        "Termo encontrado",
        "Status",
        "Trecho reportado",
    ]
    _style_header(ws, headers)
    widths = [8, 32, 8, 14, 34, 24, 18, 90]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for row in result.get("climate_policy_evidence", []):
            _write_cell(ws, row=detail_row, column=1, value=doc_idx)
            _write_cell(ws, row=detail_row, column=2, value=result["filename"])
            _write_cell(ws, row=detail_row, column=3, value=row["page"])
            _write_cell(ws, row=detail_row, column=4, value=_kind_label(row["kind"]))
            _write_cell(ws, row=detail_row, column=5, value=row["label"])
            _write_cell(ws, row=detail_row, column=6, value=row["term"])
            _write_cell(ws, row=detail_row, column=7, value=row["status"])
            _write_cell(ws, row=detail_row, column=8, value=row["snippet"])
            ws.cell(row=detail_row, column=8).alignment = Alignment(
                vertical="center", wrap_text=True
            )
            _style_detail_row(ws, detail_row, len(headers))
            detail_row += 1


def _write_climate_gaps_sheet(wb: Workbook, results: List[Dict]) -> None:
    ws = wb.create_sheet("Lacunas Climaticas")
    headers = ["N Doc.", "Arquivo", "Tipo", "Categoria", "Status"]
    _style_header(ws, headers)
    widths = [8, 32, 14, 36, 18]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for row in result.get("climate_policy_gaps", []):
            _write_cell(ws, row=detail_row, column=1, value=doc_idx)
            _write_cell(ws, row=detail_row, column=2, value=result["filename"])
            _write_cell(ws, row=detail_row, column=3, value=_kind_label(row["kind"]))
            _write_cell(ws, row=detail_row, column=4, value=row["label"])
            _write_cell(ws, row=detail_row, column=5, value=row["status"])
            _style_detail_row(ws, detail_row, len(headers))
            detail_row += 1
