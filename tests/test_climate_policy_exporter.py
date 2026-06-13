"""XLSX export tests for climate-policy outputs."""

import openpyxl
import pytest

from src.core.exporter import export_to_xlsx

pytestmark = pytest.mark.unit


def test_climate_policy_sheets_when_present(tmp_path):
    out = tmp_path / "r.xlsx"
    result = {
        "filename": "doc.pdf",
        "year": "2020",
        "president": "",
        "document": "Mensagem",
        "total_pages": 3,
        "pages_with_text": 3,
        "pages_problematic": 0,
        "ocr_pages_count": 0,
        "words_total": 100,
        "words_analytical": 80,
        "confidence": "Alto",
        "observations": "ok",
        "excluded_pages": [],
        "clim_setores_reportados": 1,
        "clim_instrumentos_reportados": 1,
        "clim_cobertura_pct": 12.5,
        "clim_intensidade": "Baixa",
        "clim_setores_dominantes": "Energia (1)",
        "clim_instrumentos_dominantes": "Planejamento (1)",
        "clim_lacunas": "Transporte",
        "climate_policy_profile": {
            "sector": {
                "direct_count": 1,
                "indirect_count": 0,
                "items": {
                    "energia": {
                        "label": "Energia",
                        "status": "reportado_direto",
                        "direct_hits": 1,
                        "indirect_hits": 0,
                        "total_hits": 1,
                    }
                },
            },
            "instrument": {
                "direct_count": 1,
                "indirect_count": 0,
                "items": {
                    "planejamento": {
                        "label": "Planejamento",
                        "status": "reportado_direto",
                        "direct_hits": 1,
                        "indirect_hits": 0,
                        "total_hits": 1,
                    }
                },
            },
        },
        "climate_policy_evidence": [
            {
                "kind": "sector",
                "item_id": "energia",
                "label": "Energia",
                "term": "energia",
                "page": 2,
                "status": "reportado_direto",
                "snippet": "Plano de energia para reduzir emissoes.",
            }
        ],
        "climate_policy_gaps": [
            {
                "kind": "sector",
                "item_id": "transporte",
                "label": "Transporte",
                "status": "nao_reportado",
            }
        ],
    }

    export_to_xlsx([result], str(out))
    wb = openpyxl.load_workbook(out)

    assert "Politica Climatica" in wb.sheetnames
    assert "Evidencias Climaticas" in wb.sheetnames
    assert "Lacunas Climaticas" in wb.sheetnames
    assert wb["Evidencias Climaticas"].cell(row=2, column=5).value == "Energia"
