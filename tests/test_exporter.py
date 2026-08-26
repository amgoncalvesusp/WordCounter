"""Unit tests for the XLSX exporter (schema-driven)."""

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

from src.core.analysis import build_column_specs, build_default_analyzers
from src.core.exporter import (
    ExcelExportError,
    _sanitize_excel_value,
    _write_cell,
    export_to_xlsx,
)

pytestmark = pytest.mark.unit


def _result(**overrides):
    base = {
        "filename": "doc.pdf",
        "year": "2020",
        "president": "Jair Bolsonaro",
        "document": "Mensagem ao Congresso Nacional",
        "total_pages": 3,
        "pages_with_text": 3,
        "pages_problematic": 0,
        "ocr_pages_count": 0,
        "words_total": 100,
        "words_analytical": 80,
        "confidence": "Alto",
        "observations": "ok",
        "excluded_pages": [],
    }
    base.update(overrides)
    return base


def test_export_creates_main_and_excluded_sheets(tmp_path):
    out = tmp_path / "r.xlsx"
    specs = build_column_specs(build_default_analyzers([], detect_sentiment=False))
    export_to_xlsx([_result()], str(out), specs)
    wb = openpyxl.load_workbook(out)
    assert "Contagem de Palavras" in wb.sheetnames
    assert "Páginas Excluídas" in wb.sheetnames


def test_export_sanitizes_illegal_xml_characters(tmp_path):
    out = tmp_path / "illegal_chars.xlsx"

    bad_text = (
        "Notas: \x0b¹ Dados consolidados em 30 de setembro de 2012. "
        "\x0c² O reconhecimento de Reservas Particulares do "
        "Patrimônio Natural (RPPN)."
    )

    result = _result(observations=bad_text)

    specs = build_column_specs(
        build_default_analyzers([], detect_sentiment=False)
    )

    export_to_xlsx([result], str(out), specs)

    wb = openpyxl.load_workbook(out)
    ws = wb["Contagem de Palavras"]

    headers = [cell.value for cell in ws[1]]
    observations_col = headers.index("Observações") + 1

    value = ws.cell(row=2, column=observations_col).value

    assert "\x0b" not in value
    assert "\x0c" not in value
    assert "¹" in value
    assert "²" in value
    assert "RPPN" in value


def test_export_writes_values_in_schema_order(tmp_path):
    out = tmp_path / "r.xlsx"
    specs = build_column_specs(build_default_analyzers([], detect_sentiment=False))
    export_to_xlsx([_result()], str(out), specs)
    ws = openpyxl.load_workbook(out)["Contagem de Palavras"]
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    assert header[0] == "Nº Doc."
    assert row[0] == 1  # doc_id auto-assigned
    assert "Jair Bolsonaro" in row


def test_export_infers_schema_without_specs(tmp_path):
    out = tmp_path / "r.xlsx"
    result = _result(
        term_results={
            "clima": {"total": 5, "analytical": 4, "exact": False, "term": "clima"}
        },
        _term_clima_total=5,
        _term_clima_analytical=4,
    )
    export_to_xlsx([result], str(out))  # no column_specs -> inferred
    ws = openpyxl.load_workbook(out)["Contagem de Palavras"]
    header = [c.value for c in ws[1]]
    assert any(h and "clima" in h for h in header)


def test_excluded_pages_written_to_second_sheet(tmp_path):
    out = tmp_path / "r.xlsx"
    result = _result(
        excluded_pages=[
            {"page_number": 1, "exclusion_reason": "provável capa", "word_count": 12}
        ]
    )
    specs = build_column_specs(build_default_analyzers([], detect_sentiment=False))
    export_to_xlsx([result], str(out), specs)
    ws = openpyxl.load_workbook(out)["Páginas Excluídas"]
    assert ws.cell(row=2, column=3).value == 1  # page number
    assert ws.cell(row=2, column=4).value == "provável capa"


def test_keyword_sheet_when_present(tmp_path):
    out = tmp_path / "r.xlsx"
    result = _result(lex_ttr=0.5, keyword_freq=[("desenvolvimento", 12), ("clima", 7)])
    export_to_xlsx([result], str(out))  # inferred schema detects text metrics
    wb = openpyxl.load_workbook(out)
    assert "Frequência de Palavras" in wb.sheetnames
    ws = wb["Frequência de Palavras"]
    assert ws.cell(row=2, column=3).value == "desenvolvimento"
    assert ws.cell(row=2, column=4).value == 12


def test_kwic_sheet_when_present(tmp_path):
    out = tmp_path / "r.xlsx"
    result = _result(
        kwic=[
            {
                "page": 2,
                "term": "clima",
                "left": "o",
                "keyword": "clima",
                "right": "mudou",
            }
        ]
    )
    specs = build_column_specs(build_default_analyzers([], detect_sentiment=False))
    export_to_xlsx([result], str(out), specs)
    wb = openpyxl.load_workbook(out)
    assert "Concordância (KWIC)" in wb.sheetnames
    ws = wb["Concordância (KWIC)"]
    assert ws.cell(row=2, column=4).value == "clima"  # termo
    assert ws.cell(row=2, column=6).value == "clima"  # ocorrência


def test_kwic_export_sanitizes_illegal_characters(tmp_path):
    out = tmp_path / "kwic_illegal.xlsx"

    result = _result(
        kwic=[
            {
                "page": 2,
                "term": "clima",
                "left": "política \x0b nacional",
                "keyword": "clima",
                "right": "mudança \x0c climática",
            }
        ]
    )

    specs = build_column_specs(
        build_default_analyzers([], detect_sentiment=False)
    )

    export_to_xlsx([result], str(out), specs)

    wb = openpyxl.load_workbook(out)
    ws = wb["Concordância (KWIC)"]

    assert "\x0b" not in ws.cell(2, 5).value
    assert "\x0c" not in ws.cell(2, 7).value
    assert "política" in ws.cell(2, 5).value
    assert "climática" in ws.cell(2, 7).value


def test_sanitize_preserves_valid_unicode_and_line_breaks():
    value = "São Paulo¹\nRPPN²\tclimática\r\nAção"

    assert _sanitize_excel_value(value) == value


def test_sanitize_removes_all_illegal_xml_control_characters():
    illegal_codepoints = [*range(0x00, 0x09), 0x0B, 0x0C, *range(0x0E, 0x20)]
    value = "|".join(chr(codepoint) for codepoint in illegal_codepoints)

    sanitized = _sanitize_excel_value(value)

    assert all(chr(codepoint) not in sanitized for codepoint in illegal_codepoints)
    assert sanitized.count(" ") == len(illegal_codepoints)


def test_export_sanitizes_bytes_values(tmp_path):
    out = tmp_path / "bytes_illegal.xlsx"
    result = _result(
        observations="Notas: \x0b¹ Dados preservados.".encode("utf-8")
    )
    specs = build_column_specs(
        build_default_analyzers([], detect_sentiment=False)
    )

    export_to_xlsx([result], str(out), specs)

    ws = openpyxl.load_workbook(out)["Contagem de Palavras"]
    observations_col = [cell.value for cell in ws[1]].index("Observações") + 1
    value = ws.cell(row=2, column=observations_col).value

    assert value == "Notas:  ¹ Dados preservados."


def test_write_cell_reports_safe_diagnostics(monkeypatch):
    ws = Workbook().active

    def fail_with_openpyxl_error(*args, **kwargs):
        raise IllegalCharacterError("texto confidencial não deve aparecer")

    monkeypatch.setattr(ws, "cell", fail_with_openpyxl_error)

    with pytest.raises(ExcelExportError) as error:
        _write_cell(ws, row=2, column=3, value="texto \x01 confidencial")

    message = str(error.value)
    assert "C2" in message
    assert "U+0001" in message
    assert "confidencial" not in message


def test_sentiment_sheet_only_when_present(tmp_path):
    out = tmp_path / "r.xlsx"
    specs = build_column_specs(build_default_analyzers([], detect_sentiment=False))
    export_to_xlsx([_result()], str(out), specs)
    assert "Sentimento (Sentenças)" not in openpyxl.load_workbook(out).sheetnames

    out2 = tmp_path / "r2.xlsx"
    result = _result(
        sent_n_sentencas=1,
        sentiment_sentences=[
            {"page": 1, "text": "Foi bom.", "compound": 0.4, "classe": "Positivo"}
        ],
    )
    export_to_xlsx([result], str(out2))
    assert "Sentimento (Sentenças)" in openpyxl.load_workbook(out2).sheetnames
